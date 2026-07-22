#!/usr/bin/env python3

import argparse
import json
from pathlib import Path
from typing import Dict, List, Any
from openpyxl import load_workbook


MODEL_ALIAS_MAP = {
    "gemma_2_2b": "M1",
    "qwen25_1_5b_instruct": "M2",
    "codegemma_2b": "M3",
    "qwen25_coder_1_5b_instruct": "M4",
    "gpt_5_1": "M5",
    "gpt_5_4": "M6",
    "gpt_5_1_codex": "M7",
    "gpt_5_3_codex": "M8",
    "gemini_3_flash_preview": "M9",
    "deepseek_chat_v32": "M10",
}


GENERATION_FILES = {
    "M1": {
        "python": "outputs/raw_generations/gemma_2_2b_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/gemma_2_2b_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/gemma_2_2b_P1_zero_shot_javascript_400_v1.jsonl",
    },
    "M2": {
        "python": "outputs/raw_generations/qwen25_1_5b_instruct_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/qwen25_1_5b_instruct_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/qwen25_1_5b_instruct_P1_zero_shot_javascript_400_v1.jsonl",
    },
    "M3": {
        "python": "outputs/raw_generations/codegemma_2b_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/codegemma_2b_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/codegemma_2b_P1_zero_shot_javascript_400_v1.jsonl",
    },
    "M4": {
        "python": "outputs/raw_generations/qwen25_coder_1_5b_instruct_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/qwen25_coder_1_5b_instruct_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/qwen25_coder_1_5b_instruct_P1_zero_shot_javascript_400_v1.jsonl",
    },
    "M5": {
        "python": "outputs/raw_generations/gpt_5_1_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/gpt_5_1_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/gpt_5_1_P1_zero_shot_javascript_400_v1.jsonl",
    },
    "M6": {
        "python": "outputs/raw_generations/gpt_5_4_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/gpt_5_4_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/gpt_5_4_P1_zero_shot_javascript_400_v1.jsonl",
    },
    "M7": {
        "python": "outputs/raw_generations/gpt_5_1_codex_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/gpt_5_1_codex_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/gpt_5_1_codex_P1_zero_shot_javascript_400_v1.jsonl",
    },
    "M8": {
        "python": "outputs/raw_generations/gpt_5_3_codex_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/gpt_5_3_codex_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/gpt_5_3_codex_P1_zero_shot_javascript_400_v1.jsonl",
    },
    "M9": {
        "python": "outputs/raw_generations/gemini_3_flash_preview_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/gemini_3_flash_preview_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/gemini_3_flash_preview_P1_zero_shot_javascript_400_v1.jsonl",
    },
    "M10": {
        "python": "outputs/raw_generations/deepseek_chat_v32_P1_zero_shot_python_400_v1.jsonl",
        "java": "outputs/raw_generations/deepseek_chat_v32_P1_zero_shot_java_400_v1.jsonl",
        "javascript": "outputs/raw_generations/deepseek_chat_v32_P1_zero_shot_javascript_400_v1.jsonl",
    },
}


SOURCE_FILES = {
    "python": "data/processed/prompted_shots/P1_zero_shot_python_400_ZS.jsonl",
    "java": "data/processed/prompted_shots/P1_zero_shot_java_400_ZS.jsonl",
    "javascript": "data/processed/prompted_shots/P1_zero_shot_javascript_400_ZS.jsonl",
}


def load_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def load_selected_ids(path: Path) -> List[str]:
    ids = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if s:
                ids.append(s)
    return ids


def build_source_lookup(root: Path) -> Dict[str, Dict[str, Any]]:
    lookup: Dict[str, Dict[str, Any]] = {}
    for lang, rel_path in SOURCE_FILES.items():
        path = root / rel_path
        rows = load_jsonl(path)
        for row in rows:
            lookup[str(row["sample_id"])] = {
                "sample_id": str(row["sample_id"]),
                "language": row.get("language", lang),
                "func_name": row.get("resolved_func_name") or row.get("func_name", ""),
                "code": row.get("code", ""),
                "reference_documentation": row.get("reference_documentation", ""),
            }
    return lookup


def build_generation_lookup(root: Path) -> Dict[str, Dict[str, str]]:
    out: Dict[str, Dict[str, str]] = {}

    for alias, lang_map in GENERATION_FILES.items():
        out[alias] = {}
        for lang, rel_path in lang_map.items():
            path = root / rel_path
            if not path.exists():
                print(f"Warning: missing generation file: {path}")
                continue
            rows = load_jsonl(path)
            for row in rows:
                sample_id = str(row["sample_id"])
                generated = row.get("generated_documentation", "")
                out[alias][sample_id] = generated
    return out


def clear_evaluation_rows(ws, start_row: int = 4, end_row: int = 5000) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, 19):
            if c == 1:
                ws.cell(r, c, f"E{r - start_row + 1:04d}")
            else:
                ws.cell(r, c, None)


def populate_sample_selection(ws, selected_ids: List[str], source_lookup: Dict[str, Dict[str, Any]]) -> None:
    start_row = 5
    for i, sample_id in enumerate(selected_ids, start=start_row):
        src = source_lookup.get(sample_id)
        if not src:
            continue
        ws.cell(i, 1, sample_id)
        ws.cell(i, 2, src["language"])
        ws.cell(i, 3, src["func_name"])
        ws.cell(i, 5, "Yes")
        ws.cell(i, 7, "zero-shot final run")
        ws.cell(i, 8, "High")


def populate_evaluation_form(
    ws,
    selected_ids: List[str],
    source_lookup: Dict[str, Dict[str, Any]],
    generation_lookup: Dict[str, Dict[str, str]],
    raters: List[str],
) -> int:
    row_num = 4
    eval_counter = 1

    aliases = list(GENERATION_FILES.keys())

    for sample_id in selected_ids:
        src = source_lookup.get(sample_id)
        if not src:
            print(f"Warning: sample_id not found in source files: {sample_id}")
            continue

        for alias in aliases:
            generated = generation_lookup.get(alias, {}).get(sample_id, "")
            if not generated:
                print(f"Warning: missing generation for {sample_id} in {alias}")
                continue

            for rater in raters:
                ws.cell(row_num, 1, f"E{eval_counter:04d}")
                ws.cell(row_num, 2, sample_id)
                ws.cell(row_num, 3, src["language"])
                ws.cell(row_num, 4, src["func_name"])
                ws.cell(row_num, 5, src["code"])
                ws.cell(row_num, 6, src["reference_documentation"])
                ws.cell(row_num, 7, alias)
                ws.cell(row_num, 8, generated)
                ws.cell(row_num, 17, rater)
                row_num += 1
                eval_counter += 1

    return eval_counter - 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--project_root", default=".", help="Root of dissertation-llm-docs")
    parser.add_argument("--workbook", required=True, help="Path to Excel workbook")
    parser.add_argument("--selected_ids", required=True, help="TXT file with one sample_id per line")
    parser.add_argument("--raters", default="R1,R2", help="Comma-separated rater IDs")
    parser.add_argument("--output", required=True, help="Output workbook path")
    args = parser.parse_args()

    root = Path(args.project_root).resolve()
    workbook_path = Path(args.workbook).resolve()
    selected_ids_path = Path(args.selected_ids).resolve()
    output_path = Path(args.output).resolve()

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    if not selected_ids_path.exists():
        raise FileNotFoundError(f"Selected IDs file not found: {selected_ids_path}")

    selected_ids = load_selected_ids(selected_ids_path)
    raters = [x.strip() for x in args.raters.split(",") if x.strip()]

    source_lookup = build_source_lookup(root)
    generation_lookup = build_generation_lookup(root)

    wb = load_workbook(workbook_path)
    ws_eval = wb["Evaluation_Form"]
    ws_select = wb["Sample_Selection"]

    clear_evaluation_rows(ws_eval)
    populate_sample_selection(ws_select, selected_ids, source_lookup)
    count = populate_evaluation_form(ws_eval, selected_ids, source_lookup, generation_lookup, raters)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Wrote populated workbook: {output_path}")
    print(f"Inserted evaluation rows: {count}")


if __name__ == "__main__":
    main()